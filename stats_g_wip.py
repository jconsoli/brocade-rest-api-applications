#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Copyright 2023, 2024 Consoli Solutions, LLC.  All rights reserved.

**License**

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

**Description**

Add statistics to Excel Workbook

Reads in the output of stats_c (which collects port statistics) and creates an Excel Workbook for each port.

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 03 Apr 2024   | Added version numbers of imported libraries.                                          |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 06 Dec 2024   | Added graphing capabilities back in.                                                  |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024 Consoli Solutions, LLC'
__date__ = '06 Dec 2024'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack@consoli-solutions.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Development'
__version__ = '4.0.3'

import sys
import os
import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcdapi.excel_util as excel_util
import brcdapi.file as brcdapi_file
import brcddb.brcddb_project as brcddb_project
import brcddb.util.util as brcddb_util
import brcddb.brcddb_common as brcddb_common
import brcddb.report.utils as report_utils
import brcddb.app_data.report_tables as rt
import brcddb.report.port as report_port
import brcddb.util.copy as brcddb_copy
import brcddb.brcddb_port as brcddb_port
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.util.search as brcddb_search
import brcddb.report.graph as report_graph
_version_d = dict(
    brcdapi_log=brcdapi_log.__version__,
    gen_util=gen_util.__version__,
    brcdapi_util=brcdapi_util.__version__,
    excel_util=excel_util.__version__,
    brcdapi_file=brcdapi_file.__version__,
    brcddb_project=brcddb_project.__version__,
    brcddb_util=brcddb_util.__version__,
    brcddb_common=brcddb_common.__version__,
    report_utils=report_utils.__version__,
    rt=rt.__version__,
    report_port=report_port.__version__,
    brcddb_copy=brcddb_copy.__version__,
    brcddb_port=brcddb_port.__version__,
    brcddb_fabric=brcddb_fabric.__version__,
    brcddb_search=brcddb_search.__version__,
    report_graph=report_graph.__version__,
)

_DOC_STRING = False  # Should always be False. Prohibits any code execution. Only useful for building documentation
# _STAND_ALONE: True: Executes as a standalone module taking input from the command line. False: Does not automatically
# execute. This is useful when importing this module into another module that calls psuedo_main().
_STAND_ALONE = True  # See note above

_graph_types = [str(k) for k in report_graph.chart_types.keys()]
_buf = 'Optional. Specifies the graph type. Valid chart types are: ' + ', '.join(_graph_types) + '. Default: line. '
_buf += 'Only used when -gp or -gs options are specified.'
_input_d = dict(
    r=dict(r=False, _r=True,
           h='Required unless -eh is specified. Report name. ".xlsx" is automatically appended.'),
    i=dict(r=False, _r=True,
           h='Required unless -eh is specified. Name of data input file. This must be the output file, -o, from '
             'stats_c.py. ".json" is automatically appended'),
    gp=dict(r=False,
            h='Optional. Creates worksheets with graphs of one or more statistical counters for specific ports. '
              'Re-run with -eh for additional help.'),
    gs=dict(r=False,
            h='Optional. Creates worksheets with graphs for a specific statistical counter. This is '
              'essentially the reverse of the -gp option. Re-run with -eh for additional help.'),
    eh=dict(r=False, t='bool', d=False,
            h='Optional. Displays extended help and exits. No other processing is performed.'),
)
_input_d.update(gen_util.parseargs_log_d.copy())

_eh_l = [
    '',
    'The -gp and -gs options create worksheets with graphs. Both options are',
    'followed with a CSV list of parameters that define the graphs. Multiple',
    'graphs can be created by separating the CSV lists with a semi-colon.',
    '',
    '**Graph by Ports, -gp**',
    '',
    'Graphs multiple statistics for a single port. Parameters that follow are:',
    '',
    '+----------+---------------------------------------------------------------+',
    '| Position | Description                                                   |',
    '+==========+===============================================================+',
    '|   0      | Port number.                                                  |',
    '+----------+---------------------------------------------------------------+',
    '|   1      | Graph type. See note 2.                                       |',
    '+----------+---------------------------------------------------------------+',
    '|   2      | A statistical parameters in                                   |',
    '|          | brocade-interface/fibrechannel-statistics. Only               |',
    '|          | the final leaf should be used.                                |',
    '+----------+---------------------------------------------------------------+',
    '|   n      | Add as many additional parameters as you wish.                |',
    '+----------+---------------------------------------------------------------+',
    '',
    'Example: Plot Tx (out-frame) and Rx (in-frame) for port 3/14 on a single',
    'line graph and CRC errors (in-crc-errors) for port 4/7 on another worksheet',
    '',
    '-gp "3/14,line,in-frames,out-frames;4/7,line,in-crc-errors"',
    '',
    '**Graph Statistics, -gs**',
    '',
    'Graphs multiple ports for a single statistic. Parameters that follow are:',
    '',
    '+----------+---------------------------------------------------------------+',
    '| Position | Description                                                   |',
    '+==========+===============================================================+',
    '|   0      | A statistical parameters in                                   |',
    '|          | brocade-interface/fibrechannel-statistics. Only the final     |',
    '|          | leaf should be used.                                          |',
    '+----------+---------------------------------------------------------------+',
    '|   1      | Graph type. See note 2.                                       |',
    '+----------+---------------------------------------------------------------+',
    '|   2      | Defines the ports to display as follows:                      |',
    '|          |                                                               |',
    '|          | E-Port Filters the list of ports resulting from other port    |',
    '|          |        parameters to just E-Ports. If no other port           |',
    '|          |        are specified, then all E-Ports in the switch are used.',
    '|          | F-Port Same as E-Port but for F-Ports.                        |',
    '|          | s/p    Port number. If "s" is omitted, "0" is assumed for     |',
    '|          |        fixed port switches. "s" and "p" may also be ranges.   |',
    '|          |        If a port in a range does not exist, it is ignored.    |',
    '|          | bot-x  Ports with the lowest total sum of the values for the  |',
    '|          |        statistic in the sample period.                        |',
    '|          | peak-x Ports with the highest peak value for the statistic.   |',
    '|          | top-x  Ports with the highest total sum of the values for the |',
    '|          |        statistic in the sample period.                        |',
    '|          |                                                               |',
    '|          | x is an integer. If x exceeds the number of ports in the      |',
    '|          | switch, the maximum number of ports in the switch are graphed.|',
    '+----------+---------------------------------------------------------------+',
    '|   n      | Additional port definitions.                                  |',
    '+----------+---------------------------------------------------------------+',
    '',
    'Example: Plot a line graph for in-frame for all ports in any switch type',
    '',
    '-gs "in-frame,line,0-12/0-63"',
    '',
    'Example: Same as above, but limited to just F-Ports.',
    '',
    '-gs "in-frame,line,F-Port,0-12/0-63"',
    '',
    'Example: Plot line graphs for in-frame and out-frame for the 5 ports with',
    '         the highest average.'
    '',
    '-gs "in-frame,line,avg-5;out-frame,line,avg-5"',
    '',
    '**Notes**',
    '',
    '1  Quotation marks around -gs and -gp may not be required; however, it',
    '   is recommended because some symbols may be interpreted as parameter',
    '   separators.',
    '2  Graph types are: ' + ', '.join(_graph_types),
    '3  The port types and graph types are not case sensitive.'
    ]

_invalid_parm_ref = 'Invalid display parameter: '
_invalid_port_ref = 'Invalid port reference: '
_sheet_map = dict()  # key: port number, value: openpyxl sheet

_port_stats = (
    'fibrechannel/average-transmit-frame-size',
    'fibrechannel/average-receive-frame-size',
    'fibrechannel/average-transmit-buffer-usage',
    'fibrechannel/average-receive-buffer-usage'
    # 'fibrechannel/current-buffer-usage',
    # 'fibrechannel/recommended-buffers',
    # 'fibrechannel/chip-buffers-available',
) + rt.Port.port_stats1_tbl


class InputError(Exception):
    pass


def _ports(proj_obj, port):
    global _invalid_port_ref

    r = list()
    switch_obj = proj_obj.r_switch_obj(proj_obj.r_get('base_switch_wwn'))
    s, p = gen_util.slot_port(port)
    if s is None or p is None:
        if gen_util.is_valid_zone_name(port):
            fab_obj = switch_obj.r_fabric_obj()
            ml = list()  # Fill this with login WWNs
            # Remember that a zone name can't be the same as an alias name so one of these loops will do nothing
            for obj in fab_obj.r_zones_for_wwn(port):  # Zone list for alias. Yes, port can be a WWN or an alias
                l = obj.r_members() + obj.r_pmembers()
                for mem in l:
                    if gen_util.is_wwn(mem):
                        ml.append(mem)
                    else:
                        obj = fab_obj.r_alias_obj(mem)
                        if obj is not None:
                            ml.extend(obj.r_members())
            obj = fab_obj.r_alias_obj(port)  # Aliases
            if obj is not None:
                ml.extend(obj.r_members())

            # Now get all the port numbers associated with the WWNs in ml
            switch_wwn = switch_obj.r_obj_key()
            for login_obj in [fab_obj.r_login_obj(mem) for mem in ml if fab_obj.r_login_obj(mem) is not None]:
                port_obj = login_obj.r_port_obj()
                if port_obj is not None and port_obj.r_switch_key() == switch_wwn:
                    # When written, login objects were only returned on the switch where the login occurred so
                    # switch_wwn has to match port_obj.r_switch_key() so this test is future proofing
                    r.append(port_obj.r_obj_key())
        else:
            brcdapi_log.log(_invalid_port_ref + port, echo=True)
    else:
        if switch_obj.r_port_obj(port) is None:
            brcdapi_log.log('Port ' + port + ' not found', echo=True)
        else:
            return list(port)
    return list()  # If we get here, something went wrong


# Case methods for _get_ports(). See _port_match
def _e_ports(switch_obj):
    return [p.r_obj_key() for p in brcddb_search.match_test(switch_obj.r_port_objects(), brcddb_search.e_ports)]


def _f_ports(switch_obj):
    return [p.r_obj_key() for p in brcddb_search.match_test(switch_obj.r_port_objects(), brcddb_search.f_ports)]


def _all_ports(switch_obj):
    return switch_obj.r_port_keys()


_port_match = {  # Used in _get_ports()
    'e': _e_ports,
    'E': _e_ports,
    'f': _f_ports,
    'F': _f_ports,
    '*': _all_ports,
}


def _get_ports(switch_obj, port_type):
    """Returns a list of port objects matching the user input.

    :param switch_obj: Switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param port_type: E, F, or *
    :type port_type: str
    :return: Port objects matching the port type
    :rtype: list
    """
    global _port_match

    r = list()
    # for port in gen_util.convert_to_list(ports.split(',')):
    #     if port in _port_match:
    #         r.extend(_port_match.get(port)(switch_obj))
    #     else:
    #         r.extend(_ports(switch_obj, port))
    #
    # r.extend(r)
    return gen_util.remove_duplicates(r)


def _get_parameters(parms):
    """Returns a list of Excel column letters matching the list of parms

    :param parms: Parameters to plot as passed in from the command line with the -p option
    :type parms: None, str, list
    :return: List of columns matching the parameters
    :rtype: list
    """
    global _invalid_parm_ref

    # Build a header to KPI reference table
    r_map = dict()
    for k, d in rt.Port.port_display_tbl.items():
        if 'fibrechannel-statistics/' in k:
            v = d.get('d')
            if v is not None:
                r_map.update({v: k})

    # Now figure out what to return
    r = list()
    for p in parms.split(','):
        v = r_map.get(p)
        if v is None:
            brcdapi_log.log(_invalid_parm_ref + p, echo=True)
        else:
            r.append(v)
    return r


def _add_ports(wb, tc_page, t_content, start_i, switch_obj):
    """Add the individual port pages to the workbook

    :param wb: Excel workbook object
    :type wb: Workbook object
    :param tc_page: Name of table of contents page
    :type tc_page: str
    :param t_content: Table of contents
    :type t_content: list
    :param start_i: Starting index (where first port goes)
    :param switch_obj: Base switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :rtype: None
    """
    global _sheet_map, _port_stats

    sheet_index = start_i
    proj_obj = switch_obj.r_project_obj()
    switch_obj_l = [proj_obj.r_switch_obj(wwn) for wwn in proj_obj.r_get('switch_list')]
    for port_obj in brcddb_util.sort_ports(switch_obj.r_port_objects()):

        # Create the port page
        port_num = port_obj.r_obj_key()
        sname = port_num.replace('/', '_')
        brcdapi_log.log('Processing port: ' + port_num, echo=True)
        port_list = [obj.r_port_obj(port_num) for obj in switch_obj_l]
        sheet = report_port.port_page(wb, '#' + tc_page + '!A1', sname, sheet_index, 'Port: ' + port_num, port_list,
                                      _port_stats, rt.Port.port_display_tbl, False)
        _sheet_map.update({port_num: sheet})

        # Add the port page to the table of contents
        t_content.append(dict(new_row=False, font='link', align='wrap', hyper='#' + sname + '!A1', disp=port_num))
        t_content.append(dict(new_row=False, font='std', align='wrap', disp=port_obj.r_port_name()))
        t_content.append(dict(new_row=False, font='std', align='wrap', disp=port_obj.c_login_type()))
        v = port_obj.r_get('fibrechannel/operational-status')
        try:
            buf = brcddb_common.port_conversion_tbl['fibrechannel/operational-status'][v]
        except KeyError:
            buf = 'Unknown'
        t_content.append(dict(new_row=False, font='std', align='wrap', disp=buf))
        t_content.append(dict(font='std', align='wrap', disp=brcddb_port.port_best_desc(port_obj)))
        sheet_index += 1


def _add_graphs(wb, tc_page, t_content, start_i, base_switch_obj, graph_list):
    """Add the individual port pages to the workbook

    Note: If there is a way to add multiple lines to a graph that aren't in neighboring columns, I haven't figured out
    how. So what we do here is add a worksheet that is a copy of the data from the different ports and or statistics
    and add it to the end. If there was just one item to plot, I wouldn't have to add another sheet but this algorithm
    is simple. It just creates a worksheet and adds columns as necessary, even if it's just one.

    :param wb: Excel workbook object
    :type wb: openpyxl.workbook.workbook.Workbook
    :param tc_page: Name of table of contents page
    :type tc_page: str
    :param t_content: Table of contents
    :type t_content: list
    :param start_i: Starting index (where first port goes)
    :param base_switch_obj: Base switch object
    :type base_switch_obj: brcddb.classes.switch.SwitchObj
    :param graph_list: List of graph dictionaries, see graph in _write_report() for details
    :type graph_list: list, tuple
    :return: List of warnings or errors.
    :rtype: list
    """
    global _sheet_map

    graph_num, sheet_index, ml = 0, start_i, list()

    proj_obj = base_switch_obj.r_project_obj()
    switch_obj_l = [proj_obj.r_switch_obj(wwn) for wwn in proj_obj.r_get('switch_list', list())]
    if len(switch_obj_l) < 2:
        brcdapi_log.log('Nothing to graph. No data collected.', echo=True)
        return ml

    for graph_d in graph_list:

        # Create the graph data page and figure out the common graphing request
        sname = 'graph_' + str(graph_num)
        brcdapi_log.log('Processing graph for: ' + sname, echo=True)
        data_sheet = wb.create_sheet(title=sname + '_data')
        data_sheet.page_setup.paperSize = data_sheet.PAPERSIZE_LETTER
        data_sheet.page_setup.orientation = data_sheet.ORIENTATION_LANDSCAPE
        port, stat = graph_d.get('port'), graph_d.get('stat')
        title, last_time, y_name = '', '', 'Programming error. Neither port or stat specified.'

        if port is not None:

            # Set up the title and chart references
            switch_obj = switch_obj_l[0]
            port_obj = switch_obj.r_port_obj(port)
            if port_obj is None:
                ml.append('Statistics for port ' + port + ' were not collected. Skipping.')
                continue
            col_ref = port
            port_page = '=' + port.replace('/', '_') + '!'

            # Find all the time stamps, reference sheets, and column references for the data reference sheet
            rd = dict()
            sheet = _sheet_map[port]
            data_sheet['A1'] = 'Time'  # Column header for the time stamp
            col = 2
            for stat in gen_util.convert_to_list(graph_d.get('parms')):
                try:
                    stat_ref = stat if rt.Port.port_display_tbl['fibrechannel-statistics/' + stat]['d'] is None else \
                        rt.Port.port_display_tbl['fibrechannel-statistics/' + stat]['d']
                except (ValueError, TypeError):
                    stat_ref = stat
                cell = excel_util.cell_match_val(sheet, stat_ref, None, 2, 1)
                if cell is None:
                    ml.append('Could not find statistical count ' + stat + ' for port ' + port + '. Skipping')
                    continue
                ref_col = column_index_from_string(coordinate_from_string(cell)[0])
                rd.update({stat: ref_col})
                data_sheet[xl.get_column_letter(col) + '1'] = port_page + cell
                col += 1
            max_col = len(rd.keys())

            # Add the time stamp
            x = port_obj.r_get(brcdapi_util.stats_time)
            if x is None:
                ml.append('Invalid sample for port ' + port + '. Skipping.')
                break
            title = 'Statistics for port ' + port + 'beginning: ' + datetime.datetime.fromtimestamp(x).strftime(
                '%d %b %Y, %H:%M:%S')
            row = 2
            for port_obj in [obj.r_port_obj(port) for obj in switch_obj_l]:
                x = None if port_obj is None else port_obj.r_get(brcdapi_util.stats_time)
                if x is None:
                    buf = 'Port ' + port + ' appears to have gone off line after ' + str(last_time)
                    buf += '. Switch: ' + port_obj.r_switch_obj().r_obj_key()
                    brcdapi_log.log(buf, echo=True)
                    break
                last_time = x
                data_sheet['A' + str(row)] = datetime.datetime.fromtimestamp(x).strftime('%H:%M:%S')
                col = 2
                for ref_col in rd.values():
                    data_sheet[xl.get_column_letter(col) + str(row)] = port_page + xl.get_column_letter(ref_col) +\
                                                                       str(row + 1)
                    col += 1
                row += 1

        if stat is not None:

            # Figure out the title and graph Y axis title
            y_name = stat.split('/').pop()
            try:
                col_ref = y_name if rt.Port.port_display_tbl[stat]['d'] is None else rt.Port.port_display_tbl[stat]['d']
            except (ValueError, TypeError) as e:
                col_ref = y_name

            # Find all the time stamps, reference sheets, and columns
            tl, rl = list(), list()
            for port in gen_util.convert_to_list(graph_d.get('parms')):
                port_obj = switch_obj_l[0].r_port_obj(port)
                if port_obj is None:
                    ml.append('Could not find port ' + port)
                    continue
                sheet = _sheet_map[port]
                cell = excel_util.cell_match_val(sheet, col_ref, None, 2, 1)
                if cell is None:
                    ml.append('Could not find column for port ' + port + ', statistic ' + stat)
                    continue
                rl.append(dict(sheet=sheet, port=port, name=port.replace('/', '_'),
                               col=column_index_from_string(coordinate_from_string(cell)[0])))
                if len(tl) == 0:
                    try:
                        x = switch_obj_l[0].r_port_obj(port).r_get(brcdapi_util.stats_time)
                    except (ValueError, TypeError):
                        x = None
                    if x is None:
                        ml.append('Invalid sample for port ' + port + '. Skipping')
                        break
                    title = 'Statistics beginning: ' + datetime.datetime.fromtimestamp(x).strftime('%d %b %Y, %H:%M:%S')
                    for port_obj in [obj.r_port_obj(port) for obj in switch_obj_l]:
                        x = port_obj.r_get(brcdapi_util.stats_time)
                        if x is None:
                            buf = 'Port ' + port + ' appears to have gone off line after ' + str(last_time)
                            buf += '. Switch: ' + port_obj.r_switch_obj().r_obj_key()
                            brcdapi_log.log(buf, echo=True)
                            break
                        last_time = x
                        tl.append(datetime.datetime.fromtimestamp(x).strftime('%H:%M:%S'))

            # Add all the data references
            max_col = len(rl)
            if max_col > 0:
                data_sheet['A1'] = 'Time'  # Column header for the time stamp
                for row in range(0, len(tl)):  # Fill in the time stamp.
                    data_sheet['A' + str(row + 2)] = tl[row]
                for col in range(0, len(rl)):  # Now add the port column headers
                    data_sheet[xl.get_column_letter(col+2) + '1'] = "'" + rl[col]['port']
                for row in range(0, len(switch_obj_l)):  # Now add all the statistical data for the ports
                    col = 2
                    for ref in rl:  # One for each port
                        data_sheet[xl.get_column_letter(col) + str(row + 2)] = \
                            '=' + ref['name'] + '!' + xl.get_column_letter(ref['col']) + str(row + 3)
                        col += 1

        # Create the Worksheet and add it to the table of contents
        max_row = len(switch_obj_l) + 1
        report_graph.graph(wb, '#' + tc_page + '!A1', sname, sheet_index,
                           dict(sheet=data_sheet,
                                title=title,
                                type=graph_d['type'],
                                x=dict(title='Time', min_col=1, min_row=1, max_row=max_row),
                                y=dict(title=col_ref, min_col=2, max_col=max_col + 1, min_row=1, max_row=max_row)))
        t_content.append(dict(merge=4, font='link', align='wrap', hyper='#' + sname + '!A1', disp=col_ref))

        sheet_index += 1
        graph_num += 1

    return ml


def _graphs(base_switch_obj, params_d):
    """Parses the graphing information from the command line into a list of machine-readable dictionaries as follows:

    +---------------+-----------------------------------------------------------------------------------------------+
    | key           | Description                                                                                   |
    +===============+===============================================================================================+
    | port          | Only present if -gp was entered on the command line. This is the port number in s/p notation  |
    |               | to plot.                                                                                      |
    +---------------+-----------------------------------------------------------------------------------------------+
    | params        | For -gp, this is the list of statistics for the port to be plotted. For -gs this is the list  |
    |               | of ports to be plotted.                                                                       |
    +---------------+-----------------------------------------------------------------------------------------------+
    | stat          | Only present if -gs was entered on the command line. This is the fibrechannel-statistics to   |
    |               | plot.                                                                                         |
    +---------------+-----------------------------------------------------------------------------------------------+
    | type          | Graph type. See brcddb.report.graph.chart_types.                                              |
    +---------------+-----------------------------------------------------------------------------------------------+
    | _ports        | Only used by this module and only used with the -gs option. This is the dynamic port          |
    |               | indicator: top-x, bot-x, peak-x, or None                                                      |
    +---------------+-----------------------------------------------------------------------------------------------+
    | _port_types   | Only used by this module and only used with the -gs option. This is the port type filter:     |
    |               | E-Port, F-Port, or None.                                                                      |
    +---------------+-----------------------------------------------------------------------------------------------+

    :param base_switch_obj: First switch object with list of ports
    :type base_switch_obj: brcddb.classes.switch.SwitchObj
    :param params_d: Dictionary of conditioned command line input. See _get_input() for details
    :type params_d: dict
    :return graphs: List of dictionaries that define the graphs. See description of graph in _write_report() for details
    :rtype graphs: list
    :return messages: Error and warning messages
    :rtype messages: list
    """
    global _port_match

    # Figure out the graph type and set up the return list of graphs
    ml, graphs, proj_obj = list(), list(), base_switch_obj.r_project_obj()

    # Single port, multiple statistics
    if isinstance(params_d['gp'], str):
        for buf in params_d['gp'].split(';'):
            temp_l = buf.split(',')
            if len(temp_l) > 1:
                port = temp_l.pop(0)
                if '/' not in port:
                    port = '0/' + port
                if base_switch_obj.r_port_obj(port) is None:
                    ml.append(port + ' not found. Skipping this port')
                else:
                    graphs.append(dict(port=port, parms=temp_l, type=params_d['gt']))

    # statistic, multiple ports
    if isinstance(params_d['gs'], str):
        for buf in params_d['gs'].split(';'):
            temp_l = buf.split(',')
            if len(temp_l) > 1:

                # Get a list of ports to consider
                try:
                    port_l = _port_match[temp_l[2] if len(temp_l) > 2 else '*'](base_switch_obj)
                except KeyError:
                    ml.append('Invalid port type, ' + temp_l[2] + ' in ' + buf + '. Re-run with -h for help')
                    continue

                statistic = 'fibrechannel-statistics/' + temp_l[0]
                to_graph = dict(stat=statistic, type=params_d['gs'])

                # Figure out which ports to plot. Note that the port list is based on the first sample only. If a port
                # logged in after the first sample, it is skipped. If a port logs out after the first sample, it will
                # still be in this list.
                if 'top-' in temp_l[1].lower() or 'avg-' in temp_l[1].lower():
                    range_l = temp_l[1].split('-')
                    # How many ports?
                    try:
                        num_ports = int(range_l[1])
                    except ValueError:
                        ml.append('Expected integer in ' + temp_l[1] + '. Re-run with -h for help')
                        continue

                    # Set up a dictionary for each port with the last count for the statistic. Statistics are
                    # cumulative. This is necessary to find the differences for the peaks and to determine total count
                    # for the sample period
                    last_port_d, port_d = dict(), dict()
                    for port in port_l:
                        last_port_d.update({port: base_switch_obj.r_port_obj(port).r_get(statistic)})
                        port_d.update({port: dict(port=port, avg=0, top=0)})

                    # Gather the statistics for each port in each sample. Keep in mind that multiple statistics are in
                    # multiple switches with -xx appended to the WWN which is used as the switch key. Also note that the
                    # switch objects are stored in a dict using wwn-xx as the key which is str. Sort on ASCII str is not
                    # a numeric order. Hence the switch_l.sort() list.
                    switch_l = [int(obj.r_obj_key().split('-')[1])
                                for obj in proj_obj.r_switch_objects() if '-' in obj.r_obj_key()]
                    switch_l.sort()
                    base_wwn = base_switch_obj.r_obj_key()
                    switch_obj_l = [proj_obj.r_switch_obj(base_wwn + '-' + str(key)) for key in switch_l]
                    for switch_obj in switch_obj_l:
                        for port in port_l:
                            port_obj = switch_obj.r_port_obj(port)
                            if port_obj is not None:  # Just in case the port went offline during the data collection
                                stat = port_obj.r_get(statistic)
                                stat_diff = stat - last_port_d[port]
                                last_port_d[port] = stat
                                port_d[port]['avg'] += stat
                                if stat_diff > port_d[port]['top']:
                                    port_d[port]['top'] = stat_diff

                    # Figure out what ports to graph
                    port_d_l = gen_util.sort_obj_num([d for d in port_d.values()], range_l[0], r=True)[0:num_ports]
                    # num_ports
                    print(len(port_d_l))

                else:
                    ml.append('Invalid ports to plot in ' + temp_l[1] + '. Re-run with -h for help')

            else:
                ml.append('Missing parameter in ' + buf + ' for -gs option. Re-run with -h for help')

    return graphs, ml


def _write_report(switch_obj, graph_list, ml, params_d):
    """Creates an Excel workbook with the port statistics differences. Operates off global data

    :param switch_obj: Base switch object with ports to write to report
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param graph_list: List of dictionaries as returned from _graphs()
    :type graph_list: list
    :param ml: Running list of messages to print after writing report.
    :type ml: list
    :param params_d: Dictionary of conditioned command line input. See _get_input() for details
    :type params_d: dict
    :return: Status code. See brcddb.brcddb_common.EXIT_xxxx
    :rtype: int
    """

    # Get the project and set up the workbook
    brcdapi_log.log('Generating Report: ' + params_d['r'] + '. This may take several seconds', echo=True)
    proj_obj = switch_obj.r_project_obj()
    wb = excel_util.new_report()

    # Set up the Project summary sheet with table of content
    title = 'Port Performance'
    tc_page = 'Project_Summary'
    t_content = [
        dict(new_row=False, merge=2, font='std', align='wrap', disp='Description'),
        dict(merge=2, font='std', align='wrap', disp=proj_obj.c_description()),
        dict(merge=2, new_row=False, font='std', align='wrap', disp='Data collected'),
        dict(merge=2, font='std', align='wrap', disp=proj_obj.r_date()),
    ]
    t_content_p = [
        dict(merge=4, font='std', align='wrap', disp=''),
        dict(merge=4, font='hdr_2', align='wrap', disp='Collected Port Data'),
        dict(merge=4, font='std', align='wrap', disp=''),
        dict(new_row=False, font='hdr_2', align='wrap', disp='Port'),
        dict(new_row=False, font='hdr_2', align='wrap', disp='Name'),
        dict(new_row=False, font='hdr_2', align='wrap', disp='Type'),
        dict(new_row=False, font='hdr_2', align='wrap', disp='State'),
        dict(font='hdr_2', align='wrap', disp='Description'),
    ]
    t_content_g = [
        dict(merge=4, font='std', align='wrap', disp=''),
        dict(merge=4, font='hdr_2', align='wrap', disp='Graphs'),
        dict(merge=4, font='std', align='wrap', disp=''),
    ]

    # Add the individual graphs and port sheets
    _add_ports(wb, tc_page, t_content_p, 0, switch_obj)
    ml.extend(_add_graphs(wb, tc_page, t_content_g, 0, switch_obj, graph_list))
    ml = gen_util.remove_duplicates(ml)
    t_content.extend(t_content_g + t_content_p)

    # Add the table of contents and save the report.
    report_utils.title_page(wb, None, tc_page, 0, title, t_content, (12, 22, 16, 10, 64))
    ml.append('Saving the report.')
    brcdapi_log.log(ml, echo=True)
    try:
        excel_util.save_report(wb, params_d['r'])
    except FileExistsError:
        brcdapi_log.log(['', 'A folder in ' + params_d['r'] + ' does not exist'], echo=True)
        return brcddb_common.EXIT_STATUS_INPUT_ERROR
    except PermissionError:
        brcdapi_log.log(['', 'Permission error writing ' + params_d['r'] + '. File may be open in another application.'],
                        echo=True)

    return brcddb_common.EXIT_STATUS_OK


def pseudo_main(proj_obj, base_switch_obj, params_d):
    """Basically the main(). Did it this way, so it can easily be used as a standalone module or called from another.

    :param params_d: Dictionary of conditioned command line input. See _get_input() for details
    :type params_d: dict
    :return: Exit code. See exit codes in brcddb.brcddb_common
    :rtype: int
    """
    global __version__, _input_d

    # Build the cross-reference tables.
    brcddb_util.build_login_port_map(proj_obj)  # Correlates name server logins with ports
    fab_obj = base_switch_obj.r_fabric_obj()
    if fab_obj is not None:
        brcddb_fabric.zone_analysis(base_switch_obj.r_fabric_obj())  # Determines what zones each login participates in

    graph_list, msg_list = _graphs(base_switch_obj, params_d)
    brcdapi_log.log('Writing ' + params_d['r'])
    return _write_report(base_switch_obj, graph_list, msg_list, params_d)


#####################################################
#                                                   #
#           Case methods for _get_input()           #
#                                                   #
#####################################################
def _e_port(port_l, split_param_l, filter_d):
    """Returns a list of E-Ports in base_switch_obj.

    :param port_l: Working list of port objects
    :type port_l: list
    :param split_param_l: Parameter list after parameter split, '-'
    :type split_param_l: list
    :param filter_d: Active filters - only used for bot-x, peak-s, and top-x
    :return r_port_l: port_l after applying the filter
    :rtype r_port_l: list
    """
    print()


def _get_input():
    """Parses the module load command line

    :return: Exit code. See exit codes in brcddb.brcddb_common
    :rtype: int
    """
    global __version__, _input_d, _version_d, _eh_l

    ec, graph_l, proj_obj, base_switch_obj = brcddb_common.EXIT_STATUS_OK, list(), None, None

    # Get command line input
    args_d = gen_util.get_input('Create Excel Workbook from statistics gathered with stats_c.py.', _input_d)

    # Set up logging
    brcdapi_log.open_log(folder=args_d['log'], supress=args_d['sup'], no_log=args_d['nl'], version_d=_version_d)

    # Extended Help
    if args_d['eh']:
        brcdapi_log.log(_eh_l, echo=True)
        return brcddb_common.EXIT_STATUS_OK

    # Make sure all required parameters were entered
    for key, input_d in _input_d.items():
        if input_d.get('_r', False) and args_d.get(key) is None:
            args_d[key] = 'ERROR: Missing required input.'
            ec = brcddb_common.EXIT_STATUS_INPUT_ERROR

    # Read in the previously collected data
    if ec == brcddb_common.EXIT_STATUS_OK:
        args_d['r'] = brcdapi_file.full_file_name(args_d['r'], '.xlsx')
        args_d['i'] = brcdapi_file.full_file_name(args_d['i'], '.json')
        try:
            obj = brcdapi_file.read_dump(args_d['i'])
            if obj is None:
                args_d['i'] = 'ERROR reading ' + args_d['i'] + '. Check the log for details.'
                ec = brcddb_common.EXIT_STATUS_ERROR
            else:
                proj_obj = brcddb_project.new(obj.get('_obj_key'), obj.get('_date'))
                proj_obj.s_python_version(sys.version)
                proj_obj.s_description(obj.get('_description'))
                brcddb_copy.plain_copy_to_brcddb(obj, proj_obj)
                obj.clear()
                base_switch_obj = proj_obj.r_switch_obj(proj_obj.r_get('base_switch_wwn'))
        except FileNotFoundError:
            args_d['i'] = 'ERROR. File not found: ' + args_d['i']
            ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
        except FileExistsError:
            args_d['i'] = 'ERROR. A folder in ' + args_d['i'] + ' does not exist.'
            ec = brcddb_common.EXIT_STATUS_INPUT_ERROR

    # Command line feedback
    ml = [
        '',
        os.path.basename(__file__) + ' version: ' + __version__,
        'Report, -r:      ' + args_d['r'],
        'Input file, -i:  ' + args_d['i'],
        'Port graph, -gp: ' + str(args_d.get('gp')),
        'Stat graph, -gs: ' + str(args_d.get('gs')),
        'Log, -log:       ' + str(args_d.get('log')),
        'No log, -nl:     ' + str(args_d.get('nl', False)),
        'Suppress, -sup:  ' + str(args_d.get('sup', False)),
        '',
        ]

    # Parse the graphing parameters (-gp and -gs)
    if ec == brcddb_common.EXIT_STATUS_OK:
        for key in [k for k in ('gp', 'gs') if isinstance(args_d.get(k), str)]:
            for graphs in args_d.get(key).split(';'):
                param_l = graphs.split(',')
                if len(param_l) < 3:
                    ml.append('At least 3 parameters are required for the -' + key + ' option. Only ' +
                              str(len(param_l)) + ' were given in ' + ','.join(param_l))
                    ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
                    break
                graph_d = dict(_params=list(), _filter_d=dict())
                if key == 'gp':
                    graph_d['port'] = param_l[0] if '/' in param_l[0] else '0/' + param_l[0]
                else:  # It must be 'gs'
                    graph_d['stat'] = param_l[0]
                if param_l[1].lower() in report_graph.chart_types:
                    graph_d['type'] = param_l[1].lower()
                else:
                    ml.append('Invalid chart type, ' + param_l[1] + ', for the -' + key + ' option.')
                    ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
                    break
                if key == 'gp':
                    graph_d['params'] = param_l[2:]
                else:  # It must be 'gs'
                    graph_param = ''
                    try:
                        for graph_param in param_l[2].split(','):
                            temp_l = gen_util.sp_range_to_list(graph_param)
                            if len(temp_l) == 0:
                                ml.append('Invalid port or port range for ' + graph_param + ' for option -gs.')
                                ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
                                break
                            graph_d['params'].extend(temp_l)
                    except InputError:
                        ml.append('Invalid port parameter for ' + graph_param + ' for option -gs')
                        ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
                        break

    brcdapi_log.log(ml, echo=True)
    return pseudo_main(proj_obj, base_switch_obj, args_d) if ec == brcddb_common.EXIT_STATUS_OK else ec


##################################################################
#
#                    Main Entry Point
#
###################################################################
if _DOC_STRING:
    print('_DOC_STRING is True. No processing')
    exit(0)

if _STAND_ALONE:
    _ec = _get_input()
    brcdapi_log.close_log(['', 'Processing Complete. Exit code: ' + str(_ec)], echo=True)
    exit(_ec)
